import openpyxl
import openpyxl.utils
wb = openpyxl.load_workbook('./file/13/test.xlsx')
print(type(wb))
print(wb.sheetnames)
print('--------------')

sheet = wb['Sheet1']
print(type(sheet))
print('--------------')

sheet1 = wb.active
print(sheet1.title)
print('--------------')

cell = sheet1['A2']
print(cell.value)
print(sheet1['B3'].value)
print(sheet1.cell(row=1, column=3).value)
print('--------------')

for i in range(1, 8, 2):
    print(i, sheet1.cell(i, 2).value)
print('--------------')

print(sheet1.max_row, sheet1.max_column)
print('--------------')

print(openpyxl.utils.get_column_letter(3))
print(openpyxl.utils.column_index_from_string('F'))
print('--------------')

cells = tuple(sheet1['B2':'C3'])
print(cells)
print(cells[0][0].value)
print('--------------')
for rowsObject in cells:
    for c in rowsObject:
        print(c.coordinate, c.value)
print('--------------')

print(list(sheet1.columns)[0]) #　获取第一列

wb.close()

# create new sheet
# wb2 = openpyxl.load_workbook('./file/13/test2.xlsx')
# wb2.create_sheet(index=0, title='hello')
# print(wb2.sheetnames)
# wb2.save('./file/13/test2.xlsx')